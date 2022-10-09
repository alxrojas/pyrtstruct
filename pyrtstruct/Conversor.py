
import os

import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from pydicom.multival import MultiValue

import Dicominfo
import validate


class Conversor():
    def __init__(self, path, name):
        self.path = path
        self.name = name
        extension = '.xlsx'
        if name.endswith(extension):
            pass
        else:
            name = ''.join([name, extension])

    def dicom2excel(self, path, name):
        # Function that creates DICOM contour in excelable form.
        # The Contour Data for each organ is set on different sheets.
        # The file is created in the same directory with the name name.xlsx.
        # INPUT:
        # Path -> The DICOM file's path.
        # Name -> str, with the name of the file.
        # OUTPUT:
        # Excel file and a dictionary with the same information.
        name = ''.join([name, '.xlsx'])
        if isinstance(path, str):
            agenda = Dicominfo.Dicominfo(path).read(path)
        else:
            agenda = path
        structures = {}
        workbook = xlsxwriter.Workbook(name)
        merge_format = workbook.add_format({'align': 'center'})
        for lesions in agenda:
            worksheet = workbook.add_worksheet(lesions)
            contour = {}
            for slices in range(len(agenda[lesions])):
                xyz_all = []
                length = int(len(agenda[lesions][slices])/3)
                param = agenda[lesions][slices]
                for delta in range(length):
                    x_coor = param[3*delta]
                    y_coor = param[3*delta + 1]
                    z_coor = param[3*delta + 2]
                    xyz = [x_coor, y_coor, z_coor, 1]
                    xyz_all.append(xyz)
                worksheet.merge_range(0, 4*slices, 0, 4*slices + 3,
                                      f'Contour {slices + 1}', merge_format)
                worksheet.write_row(1, 4*slices, ['x', 'y', 'z', 'auxiliar'])
                for delta in range(length):
                    worksheet.write_row(delta+2, 4*slices, xyz_all[delta])
                contour[slices] = xyz_all
            structures[lesions] = contour
        workbook.close()
        return structures

    def excel2dicom(self, path, name):
        # Function that creates DICOM contour in DICOM form.
        # The Contour Data for each organ is taken on different sheets.
        # The file is created in the same directory with the name name.dcm.
        # For the modification of DICOM file, it was needed to copy original.
        # File, but the Contour Data is modified.
        # The coordinates displaced (as list) to pydicom.multival.MultiValue.
        # INPUT:
        # Path -> The DICOM file's path.
        # Name -> str, with the name of the file.
        # OUTPUT:
        # DICOM file and a dictionary with the same information.
        os.chdir(path)
        extension = '.xlsx'
        if name.endswith(extension):
            pass
        else:
            name = ''.join([name, extension])
        wb = load_workbook(filename=name)
        names = wb.sheetnames
        nameid = []
        dicom = validate.validate(path)
        if dicom[1] == 'RTSTRUCT':
            indices = []
            for item in range(len(dicom[0].ROIContourSequence)):
                base = dicom[0].StructureSetROISequence[item]
                named = base.ROIName
                nameid.append(named)
            for ind in names:
                indices.append(nameid.index(ind))
            for ind in indices:
                df = pd.read_excel(name, sheet_name=nameid[ind])
                col = int(df.shape[1]/4)
                for counter in range(col):
                    x = list(df.iloc[1:, 4*counter].dropna())
                    y = list(df.iloc[1:, 4*counter + 1].dropna())
                    z = list(df.iloc[1:, 4*counter + 2].dropna())
                    # t = list(df.iloc[1:, 4*counter + 3].dropna())
                    array = []
                    for i in range(len(x)):
                        array.append(x[i])
                        array.append(y[i])
                        array.append(z[i])
                    array = np.transpose(array)
                    seq = dicom[0].ROIContourSequence[ind]
                    seq1 = seq.ContourSequence[counter]
                    seq1.ContourData = MultiValue(float, array)
            dicom[0].save_as('RS_new_dicom.dcm')
        else:
            print('Wrong DICOM modality. It is not RTSTRUCT')
        return dicom[0]
